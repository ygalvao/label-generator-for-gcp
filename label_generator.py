#!/usr/bin/env python3

#*************************************************************#
#                                                             #
#  Written by Yuri H. Galvao <yuri@galvao.ca>, November 2022  #
#                                                             #
#*************************************************************#

import pandas as pd
import requests as req
import subprocess as subp
import pathlib, time
from basic_functions import *
from datetime import datetime
from html import unescape
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.worksheet.page import PrintPageSetup
from google.cloud import storage
from oauth2client.service_account import ServiceAccountCredentials
from intuitlib.client import AuthClient
from intuitlib.enums import Scopes
from quickbooks import QuickBooks
from quickbooks.objects.invoice import Invoice

# Declaring some constants and variables
## For Intuit
CALLBACK_URI = 'https://developer.intuit.com/v2/OAuth2Playground/RedirectUrl'
intuit_keys = json.load(open('intuit_keys.json', 'r'))
intuit_temp_keys = json.load(open('intuit_temp_keys.json', 'r'))

## For Google
gcp_project = json.load(open('google-creds.json', 'r'))['project_id']

# Defining functions
def get_tokens(auth_client:object)->tuple:
    """
    """

    def finally_get_the_tokens(auth_code:str, realm_id:str)->str:
        ''''''

        ## Getting the tokens from Intuit server
        try:
            auth_client.get_bearer_token(auth_code, realm_id=realm_id)
        except Exception as e:
            print('Refresh your tokens! Exception:', e)
        else:
            access_t = auth_client.access_token
            refresh_t = auth_client.refresh_token

        return (access_t, refresh_t)

    uri = auth_client.get_authorization_url([Scopes.ACCOUNTING]) # Gets the uri from Intuit server
    print('Please, go to the following URL and authorize the app:\n\n', uri, '\n')

    time.sleep(8)

    ### Getting the auth code and realm id
    try:
        time.sleep(5)
        auth_code = input('Enter the auth code: ')
        realm_id = input('Enter the realm id: ')
    except Exception as e:
        logging.critical(f'Exception: {traceback.format_exc()}')
    else:
        auth_code = auth_code
        realm_id = realm_id
        tokens = finally_get_the_tokens(auth_code, realm_id)
        return (('access_token', tokens[0]), ('refresh_token', tokens[1]))

    return

def authenticate_on_intuit(args:list=args, intuit_keys:dict=intuit_keys, intuit_temp_keys:dict=intuit_temp_keys)->object:
    """"""

    uri = CALLBACK_URI
    auth_client = AuthClient(
        client_id=intuit_keys['client_id'],
        client_secret=intuit_keys['client_secret'],
        redirect_uri=uri,
        environment='sandbox',
        refresh_token=intuit_temp_keys['refresh_token'],
    )

    try:
        client = QuickBooks(
            auth_client=auth_client,
            refresh_token=auth_client.refresh_token,
            company_id=intuit_keys['company_id']
        )
    except Exception as e:
        logging.warning(f'Authentication error! Try refreshing the tokens!\n\nException: {e}\n')
        logging.info('The program will now try to refresh the tokens.')
        try:
            intuit_temp_keys = ask_for_data(get_tokens(auth_client), 'intuit_temp_keys', ask=False)
            auth_client.refresh_token = intuit_temp_keys['refresh_token']
            client = QuickBooks(
                auth_client=auth_client,
                refresh_token=auth_client.refresh_token,
                company_id=intuit_keys['company_id']
            )
        except:
            raise e

    return auth_client, client

def get_ds_from_api(main_object:object)->pd.Series:
    """
    Imports the data from all orders available in the designated API into a Pandas DataFrame.

    Arg.: a req.Response object.

    Returns: a Pandas DataFrame.
    """

    try:
        ds = pd.Series(main_object.to_dict())
        if ds.shape[0] <= 0:
            raise ValueError
    except ValueError:
        logging.warning('No invoice or order was found!\n')
        return
    else:
        try:
            ds = ds.apply(lambda x: unescape(x) if type(x) == str else x)
        except Exception as e:
            logging.error('Error when escaping HTML characters!')
            logging.critical(f'Exception: {e}')
        else:
            ds_ready = ds.copy()

            return ds_ready

def get_order_series(order_n:int)->pd.Series:
    """
    Fetches a specific invoice or order as a Pandas Series from an API, retrying up to 3 times if an exception occurs.

    Arg.:
        order_n (int): the invoice or order number to fetch.

    Returns:
        pd.Series: a Pandas Series containing the order data, or None if the function fails to fetch the data after 3 attempts.
    """

    ds = None
    i = 0
    auth_client, client = authenticate_on_intuit()
    while ds is None and i < 3:
        try:
            main_object = Invoice.choose([str(order_n)], field='DocNumber', qb=client)[0]
            ds = get_ds_from_api(main_object)
        except Exception as e:
            logging.error('''Error when fetching data from the designated API! Trying again in half second.''')
            logging.critical(f'''Exception: {repr(e)}''')
            ds = None
            time.sleep(.5)
            i += 1
        else:
            return ds

def get_products_names(order_series:pd.Series)->list:
    """
    Extracts a list of product names from a order as a Pandas Series.

    Args:
        order_series (pd.Series): a Pandas Series containing order data, with 'products_title' and 'products_name' fields.

    Returns:
        list: a list of formatted product names, combining 'products_title' and 'products_name' for each product in the order.
    """

    products_names = []
    for i, line in enumerate(order_series.Line):
        if line['LineNum'] == 0:
            continue

        new_product_name = line['Description'].strip()
        products_names.append(new_product_name)
    
    return products_names

def select_product(order_n:str, products_names:list)->list:
    """
    Allows the user to select the desired product from a list of available products.

    Arguments:
        order_n (int): the invoice or order number.
        products (List[str]): a list of available product names.

    Returns:
        List[str]: a list containing the selected product name.
    """

    print(f'These are the items for Invoice / Order # {order_n}:\n')
    print('Option # |   Product / Service')
    for i, product_name in enumerate(products_names):
        print(f'    {i+1}    | {product_name}')

    selected_item = int(input('Please, enter the option number for the product / service: '))
    selected_item = products_names[selected_item-1]
    logging.info(f'Selected item: {selected_item}')

    return [selected_item]

def get_address(
    order_series:pd.Series,
    type_:str,
    change:str,
    address:str,
    additional_info:str
    )->tuple:
    """
    Retrieves the 'from' or 'to' address based on the user's inputs and order data.

    Arguments:
        order_df_row (pd.Series): the order Series containing all the required data.
        address_type (str): the type of address to retrieve ('from' or 'to').
        change_address (str): a flag to indicate whether the user wants to change the address ('yes' or 'no').
        new_address (str): the new address provided by the user, if applicable.
        additional_info (str): additional information to be added to the address.

    Returns:
        tuple: a tuple containing the formatted address lines.
    """

    type_ = 'FROM' if type_.lower() == 'from' else 'SHIP TO'

    if type_ == 'FROM':
        if change.lower() == 'no':
            return

        keep_address = confirm(f'''Do you want to keep your company's address in the "{type_}" field (y/n)? ''')
        if keep_address:
            return

    if address == []:
        print(f'''For the "{type_}" field, enter the address (3 lines, in which the 1st line must be the company name and the 3rd line can be blank): ''')
        i = 0
        while i != 3:
            print("> ", end="")
            address_line = input()
            address.append(address_line)
            i += 1
    elif address == '':
        b_company = order_series.CustomerRef['name'].strip()
        d_company = b_company
        line_1 = d_company
        line_2 = f'''{order_series.ShipAddr['Line1']}''' if not order_series.ShipAddr['Line2'] else f'''{order_series.ShipAddr['Line1']}\n{order_series.ShipAddr['Line2']}'''
        line_3 = f'''{order_series.ShipAddr['City']}, {order_series.ShipAddr['CountrySubDivisionCode']}  {order_series.ShipAddr['PostalCode']}'''
        address = [line_1, line_2, line_3]
    else:
        address = address.splitlines()

    if type_ == 'FROM' and additional_info is None:
        additional_info = input('''Enter the phone number or just press "ENTER" to leave it blank: ''')

    if type_ == 'SHIP TO' and additional_info is None:
        additional_info = input('''Enter the "Attn." name or just press "ENTER" to leave it blank: ''')

    try:
        line_1, line_2, line_3 = address
    except ValueError:
        try:
            line_1, line_2 = address
        except Exception as e:
            logging.error(f'''Error when processing the "{type_}" address!''')
            logging.critical(f'''Exception: {e}''')
        else:
            return line_1, line_2, '', additional_info
    else:
        return line_1, line_2, line_3, additional_info

def get_job_data(
    order_series:pd.Series,
    order_n:int,
    add_job_info:str,
    package:str,
    packages_qty:int,
    qty_per_package:list
    )->dict:
    """
    Gets the job data from the DF, based on user's inputs and from the designated API.

    Argument:
        order_n: the order number.

    Returns:
        A dict of strings and lists.  
    """

    job_data = {}
    job_data['order_n'] = 'Order # ' + str(order_n)
    #job_data['po_n'] = order_series.PO_number if order_series.PO_number is not None else ''

    # try:
    #     int(job_data['po_n'][-4:])
    # except:
    #     job_data['po_n'] = 'Ordered by: ' + job_data['po_n']
    # else:
    #     job_data['po_n'] = 'P.O. #: ' + job_data['po_n']

    if add_job_info is None:
        job_data['additional_job_info'] = input('''Enter some additional information that you find relevant or just press "ENTER" to leave it blank: ''')
    else:
        job_data['additional_job_info'] = add_job_info

    if package is None:
        package = input('''Enter the type of package to be delivered (box, banner, case, stand, etc.): ''')

    job_data['package'] = [package.capitalize()] if len(package) > 2 else ['Box']

    while packages_qty == None:
        try:
            packages_qty = int(input('''Enter the quantity of packages (only numbers): '''))
        except ValueError:
            logging.error('Please, enter only the quantity of packages that will be delivered!')

    job_data['package'].append(packages_qty)

    while qty_per_package == None:
        try:
            qty_per_package = list_from_input('''Enter the quantity of items that are inside each package or just press "ENTER" to leave it blank: ''')
        except ValueError:
            logging.error('Please, enter only the quantity of items inside each package! Use comma to separate the quantities.')

    job_data['package'].append(qty_per_package)

    return job_data

def get_job_details(
    order_series:pd.Series,
    order_n:int,
    add_job_info:str,
    package:str,
    packages_qty:int,
    qty_per_package:list
    )->tuple:
    """
    Gets the job data from the designated API, based on user's inputs.
    """

    while order_n == None:
        try:
            order_n = int(input('Enter the order # (only numbers): '))
        except ValueError:
            logging.error('Please, enter only the number of the desired order!')

    job_details = get_job_data(order_series, order_n, add_job_info, package, packages_qty, qty_per_package)

    return job_details['order_n'], '', job_details['additional_job_info'], job_details['package']

def make_label(
    template:str,
    order_series:pd.Series,
    selected_item:list,
    spreadsheet:object,
    order_n:int,
    add_job_info:str,
    package:str,
    packages_qty:int,
    qty_per_package:list,
    from_address:str,
    to_address:str,
    additional_info_from:str,
    additional_info_to:str
    )->tuple:
    """
    Creates the shipping label using the provided workbook (Excel file) and user inputs.

    Arguments:
        template (str): the template number to be used for the label.
        order_series (pd.Series): the order Series containing all the required data.
        selected_item (list): the selected item.
        spreadsheet (object): the openpyxl workbook object.
        order_n (int): the order number.
        add_job_info (str): additional job information to be added to the label.
        package (str): the type of package to be delivered.
        packages_qty (int): the quantity of packages to be delivered.
        qty_per_package (list): the quantity of items inside each package.
        from_address (str): the 'from' address to be printed on the label.
        to_address (str): the 'to' address to be printed on the label.
        additional_info_from (str): additional information for the 'from' address.
        additional_info_to (str): additional information for the 'to' address.

    Returns:
        tuple: containing the modified workbook object, a status message, and the order number.
    """

    if order_series is None:
        order_series = get_order_series(order_n)
    
    order_df_row = order_series

    wb = spreadsheet

    from_address = get_address(
        order_df_row,
      'from',
      'no' if from_address in ('', ' ', None, []) else 'yes',
        from_address,
        additional_info_from
    )

    if to_address != '':
        to_address = get_address(
            order_df_row,
          'to',
          'yes',
            to_address,
            additional_info_to
        )
    else:
        to_address = get_address(
            order_df_row,
          'to',
          'no',
            to_address,
            additional_info_to
        )

    job_details = get_job_details(
        order_df_row,
        order_n,
        add_job_info,
        package,
        packages_qty,
        qty_per_package
    )

    selected_item = select_product(order_n, get_products_names(order_series)) if selected_item is None else selected_item

    qty_of_qties = len(qty_per_package)
    checked_items = int(qty_of_qties / packages_qty)

    for n in range(job_details[3][1]):
        if n != 0:
            wb.copy_worksheet(ws)

        ws = wb.worksheets[n]
        ws.page_margins.left = 0.
        ws.page_margins.right = 0.
        ws.page_margins.top = 0.
        ws.page_margins.bottom = 0.
        ws.column_dimensions['A'].width = 3. ##3.32  #2.75
        ws.column_dimensions['B'].width = 1.25
        ws.column_dimensions['C'].width = 25 ##33.25 #32. #28.89
        ws.column_dimensions['D'].width = 3.62 ##6.75 ## #37.5 #27.5
        ws.column_dimensions['E'].width = 27.85 ## 32.5
        ws.sheet_properties.outlinePr.applyStyles = True
        ws.sheet_properties.pageSetUpPr.fitToPage = False
        ws.delete_cols(6,4)
        ws.print_area = 'A1:E12'
        ws.set_printer_settings(0, orientation='landscape')
        ws.page_setup.paperHeight = '152mm'
        ws.page_setup.paperWidth = '102mm'

        if n != 0:
            logo = Image('logo_for_xlsx.png')
            ws.add_image(logo, 'E1')

        if from_address is not None:
            ws['C1'], ws['C2'], ws['C3'], ws['C4'] = from_address

        ws['C5'], ws['C6'], ws['C7'], ws['C8'] = to_address
        ws['C9'], ws['C10'], ws['C12'], package_data = job_details

        if template in ('', ' ', '1', None):
            try:
                ws['C11'] = selected_item[n]
            except:
                ws['C11'] = selected_item[0]

            ws['E8'] = f'{package_data[0]} {n+1} of {package_data[1]}  '

            try:
                ws['D11'] = f'Total qty: {sum(package_data[2])}  '
            except:
                ws['D11'] = f'Total qty: {package_data[2]}  '

            try:
                ws['D12'] = f'Qty in this {package_data[0].lower()}: {package_data[2][n]}  '
            except:
                ws['D12'] = f'Qty in this {package_data[0].lower()}: {package_data[2]}  '
        else:
            try:
                ws['C11'] = selected_item[n]
            except:
                ws['C11'] = selected_item[0]

            ws['E8'] = f'{package_data[0]} {n+1} of {package_data[1]}  '

            ws['E10'] = f'Qty in this {package_data[0].lower()}:'

            try:
                qties_info = ''                
                for i, n in enumerate(package_data[2]):
                    if i < checked_items:
                        qties_info += f'Item {i+1}: {n}\n'

                ws['E11'] = qties_info
                qties_info = ''
                package_data[2] = package_data[2][checked_items:]           
            except Exception as e:
                ws['E11'] = f'Item 1: {package_data[2]}'

    return wb, 'Finished', job_details[0].split(' ')[-1]

def upload_to_bucket(blob_name:str, path_to_file:str, bucket_name:str)->str:
    """
    Uploads a file to a specified Google Cloud Storage bucket.

    Arguments:
        blob_name (str): the name to be used for the blob (file) in the bucket.
        path_to_file (str): the local file path of the file to be uploaded.
        bucket_name (str): the name of the Google Cloud Storage bucket to upload the file to.

    Returns:
        str: a public URL to access the uploaded file in the Google Cloud Storage bucket.
    """

    # Explicitly use service account credentials by specifying the private key file.
    storage_client = storage.Client.from_service_account_json('google-creds.json')

    #print(buckets = list(storage_client.list_buckets())
    bucket = storage_client.get_bucket(bucket_name)
    blob = bucket.blob(blob_name)
    blob.upload_from_filename(path_to_file)

    os.remove(path_to_file)
    
    # Returns a public url
    return blob.public_url

def generate_pdf(file_name:str)->None:
    """
    Converts the Excel file (.xlsx) into a PDF file and uploads it to a specified Google Cloud Storage bucket.

    Arguments:
        file_name (str): the name of the Excel file to be converted.

    Returns:
        str: the public URL to access the generated PDF file in the Google Cloud Storage bucket.
    """

    cmd = f'libreoffice --headless --convert-to pdf --outdir ./output {file_name}'

    try:
        p = subp.run(cmd, shell=True)
    #convert_to_pdf = UnoConverter()
    #convert_to_pdf.convert(inpath=file_name, outpath='print_this.pdf', convert_to='pdf')
    except Exception as e:
        logging.error(e)
    else:
        file_name_pdf = file_name[:-5]+'.pdf'
    finally:
        os.remove(file_name)

    url_to_pdf = upload_to_bucket(file_name_pdf, './output/'+file_name_pdf, f'{gcp_project}-processed-labels')

    logging.info(f'PDF created!\n')
    logging.info(f'Download it here: {url_to_pdf}')

    return url_to_pdf

def output_label(
    template:str='',
    order_series:pd.Series=None,
    order_n:int=None,
    selected_item:list=None,
    add_job_info:str=None,
    package:str=None,
    packages_qty:int=None,
    qty_per_package:list=None,
    from_address:str=[],
    to_address:str=[],
    additional_info_from:str=None,
    additional_info_to:str=None
    )->tuple:
    """
    Generates a shipping label based on the given inputs and saves it as an Excel file (.xlsx). Then, converts the Excel file into a PDF file, and uploads it to a Google Cloud Storage bucket.

    Arguments:
        template (str, optional): the template number to be used for the label. Default is an empty string.
        order_series (pd.Series, optional): the order Series containing all the required data. Default is None.
        order_n (int, optional): the order number. Default is None.
        selected_item (list, optional): the selected item. Default is None.
        add_job_info (str, optional): additional job information to be added to the label. Default is None.
        package (str, optional): the type of package to be delivered. Default is None.
        packages_qty (int, optional): the quantity of packages to be delivered. Default is None.
        qty_per_package (list, optional): the quantity of items inside each package. Default is None.
        from_address (str, optional): the 'from' address to be printed on the label. Default is an empty list.
        to_address (str, optional): the 'to' address to be printed on the label. Default is an empty list.
        additional_info_from (str, optional): additional information for the 'from' address. Default is None.
        additional_info_to (str, optional): additional information for the 'to' address. Default is None.

    Returns:
        tuple: a tuple containing a status message ('Success') and the public URL to access the generated PDF file in the Google Cloud Storage bucket.
    """

    local_path = pathlib.Path().resolve().__str__() # Gets the local path

    try:
        wb = load_workbook(f'./templates/template{template}.xlsx')
    except FileNotFoundError:
        try:
            wb = load_workbook(local_path + f'\\templates\\template{template}.xlsx')
        except Exception as e:
            logging.error(f'''Path error! Check your path to "template{template}.xlsx".''')
            logging.critical(f'''Exception: {e}''')
    except Exception as e:
        logging.error('''Unknown error!''')
        logging.critical(f'''Exception: {e}''')

    wb, status, order_n = make_label(
        template,
        order_series,
        selected_item,
        wb,
        order_n,
        add_job_info,
        package,
        packages_qty,
        qty_per_package,
        from_address,
        to_address,
        additional_info_from,
        additional_info_to
        )

    if status == 'Finished':
        logging.info('Label made successfully!\n')

    now = datetime.now().strftime("%Y-%m-%d_%H:%M")
    file_name = f'final_label_-_{now}_-_Order_{order_n}.xlsx'

    wb.save(file_name)
    logging.info(f'File {file_name} saved!\n')

    link_to_pdf = generate_pdf(file_name)

    return 'Success', link_to_pdf

if __name__ == '__main__':
    output_label()
    logging.info('End of the program.\n')
